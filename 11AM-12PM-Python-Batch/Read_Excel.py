'''
Read Excel File

install library
pip install openpyxl



from openpyxl import load_workbook
workbook = load_workbook('Financial_Sample.xlsx')
sheet = workbook.active
for row in sheet.iter_rows(values_only=True):
    print(row)



jupyter notebook

pip install jupyter

'''


